import openpyxl


class Homepagedata:
    test_Homepage_data = [{"firstname":"Mariano", "email":"mmc@gmail.com", "gender":"Male"}, {"firstname":"Caro", "email":"caro@gmail.com", "gender":"Female"}]

    @staticmethod
    def getTestData(test_case_name): # We can delete static because it is an static method- We are doing this to avoid to create a method
        dict = {}
        book = openpyxl.load_workbook("C:\\Users\EE857CF\\OneDrive - EY\\Documents\\02- Udemy EY\\Selenium Webdriver with PYTHON from Scratch + Frameworks\\Data from Excel\\File.xlsx")  # Open the book
        sheet = book.active  # Control of sheet object
        for i in range(1, sheet.max_row + 1):  # To Get the rows
            if sheet.cell(row=i, column=1).value == test_case_name: # We are sending as parameter the test case name....
                for g in range(2, sheet.max_column + 1):  # To get the columns
                    # Store all information in a dictionary
                    dict[sheet.cell(row=1, column=g).value] = sheet.cell(row=i, column=g).value

        return [dict]

