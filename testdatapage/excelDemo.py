import openpyxl
book = openpyxl.load_workbook("C:\\Users\EE857CF\\OneDrive - EY\\Documents\\02- Udemy EY\\Selenium Webdriver with PYTHON from Scratch + Frameworks\\Data from Excel\\File.xlsx") # Open the book
dict = {}
sheet = book.active # Control of sheet object

cell = sheet.cell(row=1,column=2)
print(cell.value)

sheet.cell(row=2,column=2).value = "Mariano" # Write a value in a columns

print(sheet.cell(row=2,column=2).value)

print(sheet.max_row) #How many rows are in a sheet

print(sheet.max_column) #How many rows are in a sheet

print(sheet['A5'].value) # Another way of see the information

for i in range(1, sheet.max_row+1): # To Get the rows
    if sheet.cell(row=i, column=1).value == "Testcase1":
        for g in range(2, sheet.max_column+1): # To get the columns
            # Store all information in a dictionary
            dict[sheet.cell(row=1, column=g).value] = sheet.cell(row=i, column=g).value


print(dict)



